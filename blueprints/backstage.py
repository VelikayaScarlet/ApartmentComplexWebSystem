import requests
from openpyxl import *
from flask import *

from datetime import *
from blueprints.forms import *
from model import *

bp = Blueprint('backstage', __name__, url_prefix='/backstage')


@bp.route('/car_bs')
def car_bs():
    cars = CarModel.query.order_by(CarModel.vp).all()
    return render_template('e_car_bs.html', cars=cars)


@bp.route('/temp_vehicle_bs')
def temp_vehicle_bs():
    temp_vehicles = TempVehicleModel.query.order_by(TempVehicleModel.id).all()
    return render_template('e_vehicle_bs.html', temp_vehicles=temp_vehicles)


@bp.route('/resident_bs')
def resident_bs():
    residents = ResidentModel.query.order_by(ResidentModel.id).all()
    return render_template('e_resident_bs.html', residents=residents)


@bp.route('/visitor_bs')
def visitor_bs():
    visitors = VisitorModel.query.order_by(VisitorModel.visit_time.desc()).all()
    today = date.today()
    f_t_month = today.strftime("%m")
    f_t_year = today.strftime("%y")
    f_today = today.strftime("%d")
    monday = date.today()
    sunday = date.today()
    one_day = timedelta(days=1)
    while monday.weekday() != 0:
        monday -= one_day
    while sunday.weekday() != 6:
        sunday += one_day
    md = monday.strftime("%y%m%d")
    mon = tue = wed = thu = fri = sat = sun = 0
    for v in visitors:
        vt = v.visit_time.strftime("%y%m%d")
        if vt > md:
            if v.visit_time.weekday() == 0:
                mon += 1
            if v.visit_time.weekday() == 1:
                tue += 1
            if v.visit_time.weekday() == 2:
                wed += 1
            if v.visit_time.weekday() == 3:
                thu += 1
            if v.visit_time.weekday() == 4:
                fri += 1
            if v.visit_time.weekday() == 5:
                sat += 1
            if v.visit_time.weekday() == 6:
                sun += 1
        else:
            break
    date_data = [mon, tue, wed, thu, fri, sat, sun]
    weekdays = []
    for i in range(0, 7):
        weekdays.append(int(monday.strftime('%m%d')))
        monday += one_day
    a = b = c = d = e = f = 0
    f_today = int(f_today)
    for v in visitors:
        vty = int(v.visit_time.strftime('%y'))
        vtm = int(v.visit_time.strftime('%m'))
        vtd = int(v.visit_time.strftime("%d"))
        vth = int(v.visit_time.strftime("%H"))
        if vtd == f_today and vtm == f_t_month and vty == f_t_year:
            if 0 <= vth < 4:
                a += 1
            if 4 <= vth < 8:
                b += 1
            if 8 <= vth < 12:
                c += 1
            if 12 <= vth < 16:
                d += 1
            if 16 <= vth < 20:
                e += 1
            if 20 <= vth < 24:
                f += 1
    today_list = [a, b, c, d, e, f]
    return render_template('e_visitor_bs.html', visitors=visitors, date_data=date_data, weekdays=weekdays,
                           today_list=today_list)


@bp.route('/resident_profile/<resident_id>')
def resident_profile(resident_id):
    resident = ResidentModel.query.get(resident_id)
    family = ResidentModel.query.filter_by(building=resident.building, room=resident.room).all()
    records = PARecordModel.query.filter_by(person_type='住户', name=resident.name, person_id=resident.id).order_by(PARecordModel.time.desc()).limit(10).all()
    return render_template('f_resident_profile.html', resident=resident, family=family, records=records)


@bp.route('/resident_search', methods=['POST', 'GET'])
def resident_search():
    ctt = request.form.get('ctt')
    print(ctt)  # 这个print 我不知道有什么用,但有了它，程序就正常运行
    if ctt is None:
        ctt = ''
    residents = ResidentModel.query.order_by(ResidentModel.id).all()
    resident_list = []
    for r in residents:
        if ctt in str(r.id) or ctt in r.name or ctt in str(r.phone_num) or ctt in r.workplace or ctt in str(
                r.building) or ctt in str(r.room):
            resident_list.append(r)
    return render_template('e_resident_bs.html', residents=resident_list)


@bp.route('/visitor_search', methods=['POST', 'GET'])
def visitor_search():
    ctt = request.form.get('ctt')
    print(ctt)  # 这个print 我不知道有什么用,但有了它，程序就正常运行
    if ctt is None:
        ctt = ''
    visitors = VisitorModel.query.order_by(VisitorModel.id).all()
    visitor_list = []
    for r in visitors:
        if ctt in str(r.id) or ctt in r.name or ctt in str(r.phone_num) or ctt in r.reason or ctt in str(r.visit_time):
            visitor_list.append(r)
    return render_template('e_visitor_bs.html', visitors=visitor_list)


@bp.route('/visitor_le/<visitor_id>', methods=['POST', 'GET'])
def visitor_le(visitor_id):
    visitor = VisitorModel.query.filter_by(id=visitor_id).first()

    if visitor.le_type == '没走':
        visitor.le_type = '走了'
        le_type = '出'
    else:
        visitor.le_type = '没走'
        le_type = '入'
    record = PARecordModel(person_id=visitor.id,
                           name=visitor.name,
                           person_type='访客',
                           le_type=le_type,
                           time=datetime.now())
    db.session.add(record)
    db.session.commit()
    return redirect(url_for('backstage.visitor_bs'))


@bp.route('/car_search', methods=['POST', 'GET'])
def car_search():
    ctt = request.form.get('ctt')
    print(ctt)  # 这个print 我不知道有什么用,但有了它，程序就正常运行
    if ctt is None:
        ctt = ''
    cars = CarModel.query.order_by(CarModel.id).all()
    car_list = []
    for r in cars:
        if ctt in str(r.id) or ctt in r.vp or ctt in str(r.owner_id) or ctt in r.owner.name:
            car_list.append(r)
    return render_template('e_car_bs.html', cars=car_list)


@bp.route('/temp_vehicle_search', methods=['POST', 'GET'])
def temp_vehicle_search():
    ctt = request.form.get('ctt')
    print(ctt)  # 这个print 我不知道有什么用,但有了它，程序就正常运行
    if ctt is None:
        ctt = ''
    tvs = TempVehicleModel.query.order_by(TempVehicleModel.id).all()
    tv_list = []
    for r in tvs:
        if ctt in str(r.id) or ctt in r.vp or ctt in str(r.owner_id) or ctt in r.owner.name:
            tv_list.append(r)
    return render_template('e_vehicle_bs.html', temp_vehicles=tv_list)


@bp.route('/resident_delete/<resident_id>')
def resident_delete(resident_id):
    resident = ResidentModel.query.get(resident_id)
    cars = CarModel.query.filter_by(owner_id=resident.id)
    for car in cars:
        db.session.delete(car)
    db.session.delete(resident)
    db.session.commit()
    return redirect(url_for('backstage.resident_bs'))


@bp.route('/visitor_edit/<visitor_id>', methods=['POST', 'GET'])
def visitor_edit(visitor_id):
    if request.method == 'GET':
        print('get')
    else:
        form = EditVisitorForm(request.form)
        id = visitor_id
        name = form.name.data
        phone_num = form.phone_num.data
        reason = form.reason.data
        visitor = db.session.query(VisitorModel).filter_by(id=id).first()
        visitor.name = name
        visitor.phone_num = phone_num
        visitor.reason = reason
        db.session.commit()
    # return render_template('e_resident_bs.html', visitors=visitors)
    return redirect(url_for('backstage.visitor_bs'))


@bp.route('/resident_edit/<resident_id>', methods=['POST', 'GET'])
def resident_edit(resident_id):
    if request.method == 'GET':
        print('get')
    else:
        print('post')
        form = EditResidentForm(request.form)
        id = resident_id
        name = form.name.data
        phone_num = form.phone_num.data
        wp = form.workplace.data
        resident = db.session.query(ResidentModel).filter_by(id=id).first()
        resident.name = name
        resident.phone_num = phone_num
        resident.workplace = wp
        db.session.commit()
    # return render_template('e_resident_bs.html', residents=residents)
    return redirect(url_for('backstage.resident_bs'))


@bp.route('/car_edit<car_id>', methods=['POST', 'GET'])
def car_edit(car_id):
    if request.method == 'GET':
        print('get')
    else:
        print('post')
        form = EditCarForm(request.form)
        id = car_id
        vp = form.vp.data
        # owner_id = form.owner_id.data
        car = db.session.query(CarModel).filter_by(id=id).first()
        car.vp = vp
        db.session.commit()
    # return render_template('e_car_bs.html', cars=cars)
    return redirect(url_for('backstage.car_bs'))


@bp.route('temp_vehicle_edit/<tv_id>', methods=['POST', 'GET'])
def temp_vehicle_edit(tv_id):
    if request.method == 'GET':
        print('get')
    else:
        print('post')
        form = EditCarForm(request.form)
        id = tv_id
        vp = form.vp.data
        # owner_id = form.owner_id.data
        tv = db.session.query(TempVehicleModel).filter_by(id=id).first()
        tv.vp = vp
        db.session.commit()
    # return render_template('e_vehicle_bs.html', temp_vehicles=tvs)
    return redirect(url_for('backstage.temp_vehicle_bs'))


@bp.route('/visitor_delete/<visitor_id>')
def visitor_delete(visitor_id):
    visitor = VisitorModel.query.get(visitor_id)
    tvs = TempVehicleModel.query.filter_by(owner_id=visitor.id)
    for tv in tvs:
        db.session.delete(tv)
    db.session.delete(visitor)
    db.session.commit()
    return redirect(url_for('backstage.resident_bs'))


@bp.route('/car_delete/<car_id>')
def car_delete(car_id):
    car = CarModel.query.get(car_id)
    db.session.delete(car)
    db.session.commit()
    return redirect(url_for('backstage.car_bs'))


@bp.route('/temp_vehicle_delete/<tv_id>')
def temp_vehicle_delete(tv_id):
    tv = TempVehicleModel.query.get(tv_id)
    db.session.delete(tv)
    db.session.commit()
    return redirect(url_for('backstage.car_bs'))


@bp.route('/info')
def info():
    pass


@bp.route('/parking')
def parking():
    pass


@bp.route('/export_resident')
def export_resident():
    wb = Workbook()
    ws = wb.active
    residents = ResidentModel.query.order_by(ResidentModel.id).all()
    ws['A1'] = 'ID'
    ws['B1'] = '姓名'
    ws['C1'] = '是否为业主'
    ws['D1'] = '联系方式'
    ws['E1'] = '居住单元'
    ws['F1'] = '工作单位'
    for resident in residents:
        ih = '否'
        if resident.is_holder:
            ih = '是'
        ws.append([resident.id, resident.name, ih, resident.phone_num,
                   str(resident.building) + '#' + str(resident.room), resident.workplace])
    wb.save('C:\\Users\\79433\\Desktop\\业主信息.xlsx')
    return redirect(url_for('backstage.resident_bs'))


@bp.route('/export_visitor')
def export_visitor():
    wb = Workbook()
    ws = wb.active
    visitors = VisitorModel.query.order_by(VisitorModel.id).all()
    ws['A1'] = 'ID'
    ws['B1'] = '姓名'
    ws['C1'] = '原因'
    ws['D1'] = '联系方式'
    ws['E1'] = '访问时间'
    for visitor in visitors:
        ws.append([visitor.id, visitor.name, visitor.reason, visitor.phone_num, visitor.visit_time])
    wb.save('C:\\Users\\79433\\Desktop\\访客信息.xlsx')
    return redirect(url_for('backstage.resident_bs'))


@bp.route('/export_car')
def export_car():
    wb = Workbook()
    ws = wb.active
    cars = CarModel.query.order_by(CarModel.id).all()
    ws['A1'] = '车辆牌照'
    ws['B1'] = '车主编号'
    ws['C1'] = '车主姓名'
    ws['D1'] = '车主电话'
    for car in cars:
        ws.append([car.vp, car.owner_id, car.owner.name, car.owner.phone_num])
    wb.save('C:\\Users\\79433\\Desktop\\业主车辆信息.xlsx')
    return redirect(url_for('backstage.car_bs'))


@bp.route('/export_vehicle')
def export_vehicle():
    wb = Workbook()
    ws = wb.active
    vehicles = TempVehicleModel.query.order_by(TempVehicleModel.id).all()
    ws['A1'] = '车辆牌照'
    ws['B1'] = '车主编号'
    ws['C1'] = '车主姓名'
    ws['D1'] = '车主电话'
    for car in vehicles:
        ws.append([car.vp, car.owner_id, car.owner.name, car.owner.phone_num])
    wb.save('C:\\Users\\79433\\Desktop\\外来车辆信息.xlsx')
    return redirect(url_for('backstage.car_bs'))


@bp.route('/parecord')
def parecord():
    records = PARecordModel.query.order_by(PARecordModel.time.desc()).all()
    return render_template('e_parecord.html', records=records)


@bp.route('/export_parecord')
def export_parecord():
    wb = Workbook()
    ws = wb.active
    records = PARecordModel.query.order_by(PARecordModel.time.desc()).all()
    ws['A1'] = '记录编号'
    ws['B1'] = '人员编号'
    ws['C1'] = '人员姓名'
    ws['D1'] = '人员类型'
    ws['E1'] = '行动'
    ws['F1'] = '时间'
    for r in records:
        ws.append([r.id, r.person_id, r.name, r.person_type, r.le_type, r.time])
    wb.save('C:\\Users\\79433\\Desktop\\人员进出记录.xlsx')
    return redirect(url_for('backstage.parecord'))


@bp.route('/parecord_search', methods=['POST', 'GET'])
def parecord_search():
    ctt = request.form.get('ctt')
    print(ctt)  # 这个print 我不知道有什么用,但有了它，程序就正常运行
    if ctt is None:
        ctt = ''
    records = PARecordModel.query.order_by(PARecordModel.id).all()
    r_list = []
    for r in records:
        if ctt in str(r.id) or ctt in str(r.person_id) or ctt in r.name or ctt in r.person_type or \
                ctt in r.le_type or ctt in str(r.time):
            r_list.append(r)
    return render_template('e_varecord.html', records=r_list)


@bp.route('/varecord')
def varecord():
    records = VARecordModel.query.order_by(VARecordModel.time.desc()).all()
    return render_template('e_varecord.html', records=records)


@bp.route('/export_varecord')
def export_varecord():
    wb = Workbook()
    ws = wb.active
    records = VARecordModel.query.order_by(VARecordModel.time.desc()).all()
    ws['A1'] = '记录编号'
    ws['B1'] = '车辆编号'
    ws['C1'] = '车牌号'
    ws['D1'] = '车辆类型'
    ws['E1'] = '行动'
    ws['F1'] = '时间'
    for r in records:
        ws.append([r.id, r.vehicle_id, r.vp, r.vehicle_type, r.le_type, r.time])
    wb.save('C:\\Users\\79433\\Desktop\\车辆进出记录.xlsx')
    return redirect(url_for('backstage.parecord'))


@bp.route('/varecord_search', methods=['POST', 'GET'])
def varecord_search():
    ctt = request.form.get('ctt')
    print(ctt)  # 这个print 我不知道有什么用,但有了它，程序就正常运行
    if ctt is None:
        ctt = ''
    records = VARecordModel.query.order_by(VARecordModel.id).all()
    r_list = []
    for r in records:
        if ctt in str(r.id) or ctt in str(r.vehicle_id) or ctt in r.vp or ctt in r.vehicle_type or \
                ctt in r.le_type or ctt in str(r.time):
            r_list.append(r)
    return render_template('e_varecord.html', records=r_list)